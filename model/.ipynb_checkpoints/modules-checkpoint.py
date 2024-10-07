import pandas as pd
import numpy as np
from plotly.subplots import make_subplots
from tqdm import tqdm
import plotly.graph_objects as go
from fpdf import FPDF
import os
import openpyxl
from itertools import product
import sys

path = os.getcwd()

project_directory = os.path.dirname( os.path.abspath( path ) )

sys.path.append(project_directory + "/model")
from modelling_outil import cleaning_data, copy_formatting , highlight_above_threshold , determine_MP_set , generate_recommendation_summary

class Factory:
    def __init__(self, factory_df):
        """
        Intiate a factory object instance from the factory dataframe.
        :param self: a factory object instance
        :type self: Factory object
        :param factory_df: a dataframe of the factory 
        :type factory_df: pd.Dataframe
        """
        self.formulas_dict = {}
        self.ingest_factory_data(factory_df)
    
    def ingest_factory_data(self,factory_df):
        """
        Ingest a factory data from the factory dataframe.
        :param self: a factory object
        :type self: Factory object
        :param factory_df: a dataframe of the factory 
        :type factory_df: pd.Dataframe
        """
        factory_df_cleaned = cleaning_data(factory_df)
        formulas_set = factory_df_cleaned['N° formule'].unique().tolist()

        for formula in tqdm(formulas_set,desc="Processing formulas"):
            formula_df = factory_df_cleaned[factory_df_cleaned['N° formule'] == formula]
            formula_created = Formula(formula_df = formula_df)
            self.formulas_dict[formula] = formula_created

    def get_adjusted_formulas_ids(self):
        """
        Determine list of adjusted formulas ids.
        :param self: a factory object
        :type self: Factory object
        """
        adjusted_formulas_ids_list = []
        for formula_id in self.formulas_dict:
            if self.formulas_dict[formula_id].number_of_adjusted_batches > 0 :
                adjusted_formulas_ids_list.append(formula_id)
        return adjusted_formulas_ids_list
    
    def get_no_unique_raw_material_composition_formulas_ids(self):
        """
        Determine list of formulas ids that have non-unique raw material composition.
        :param self: a factory object
        :type self: Factory object
        """
        no_unique_raw_material_composition_formulas_ids_list = []
        for formula_id in self.formulas_dict:
            if self.formulas_dict[formula_id].no_unique_raw_material_composition == True :
                no_unique_raw_material_composition_formulas_ids_list.append(formula_id)
        return no_unique_raw_material_composition_formulas_ids_list

    def select_formulas_of_cv_higher_than_threshold(self,threshold):
        """
        Determine list of formulas ids that have coefficient of variation of at least one raw material higher than specified threshold.
        :param self: a factory object
        :type self: Factory object
        :param threshold: specified threshold
        :type threshold: float
        """
        selected_formulas = []
        for formula_id in self.formulas_dict:
            for value in self.formulas_dict[formula_id].coefficient_of_variation_of_MP.values():
                if value > threshold:
                    selected_formulas.append(formula_id)
                    break
        return selected_formulas

    def select_formulas_of_cv_lower_equal_than_threshold(self,threshold):
        """
        Determine list of formulas ids that have coefficient of variation of at least one raw material lower than or equal to specified threshold.
        :param self: a factory object
        :type self: Factory object
        :param threshold: specified threshold
        :type threshold: float
        """
        selected_formulas_ids = []
        for formula_id in self.formulas_dict:
            for value in self.formulas_dict[formula_id].coefficient_of_variation_of_MP.values():
                if value <= threshold:
                    selected_formulas_ids.append(formula_id)
        return selected_formulas_ids
                
    def plot_selected_formulas_to_pdf(self , selected_formulas_ids , filename):
        """
        Export stacked bar graph plots of the raw material composition of selected formulas ids to a pdf file.
        :param self: a factory object
        :type self: Factory object
        :param selected_formulas_ids: list of selected formulas ids to plot
        :type selected_formulas_ids: list
        :param filename: pdf filename
        :type filename: string
        """
        figures = []
        for selected_formula_id in selected_formulas_ids:
            selected_formula = self.formulas_dict[selected_formula_id]
            fig = selected_formula.plot_bar_graph()
            figures.append(fig)
        
        pdf = FPDF()
        pdf.add_page()
        for fig_index , fig in enumerate(figures,1):
            img_path = f"plot_{fig_index}.png"
            fig.write_image(img_path)

            if pdf.get_y() + 100 > pdf.h :
                pdf.add_page()
            
            pdf.image(img_path , x=10 , y=pdf.get_y() + 30 , w=170)
            pdf.set_y(pdf.get_y() + 100)

            os.remove(img_path)

        pdf.output(filename,'F')

    def plot_selected_formulas(self , selected_formulas_ids ):
        """
        Print out stacked bar graph plots of the raw material composition of selected formulas ids.
        :param self: a factory object
        :type self: Factory object
        :param selected_formulas_ids: list of selected formulas ids to plot
        :type selected_formulas_ids: list
        """
        figures = []
        for selected_formula_id in selected_formulas_ids:
            selected_formula = self.formulas_dict[selected_formula_id]
            fig = selected_formula.plot_bar_graph()
            figures.append(fig)

        for fig in figures:
            fig.show()

    def plot_selected_formulas_with_ref_to_pdf(self , selected_formulas_ids , filename):
        """
        Export stacked bar graph plots of the raw material composition ratio with respect to reference batch of the selected formulas ids to a pdf file.
        :param self: a factory object
        :type self: Factory object
        :param selected_formulas_ids: list of selected formulas ids to plot
        :type selected_formulas_ids: list
        :param filename: pdf filename
        :type filename: string
        """
        figures = []
        for selected_formula_id in selected_formulas_ids:
            selected_formula = self.formulas_dict[selected_formula_id]
            fig = selected_formula.plot_bar_graph_with_ref()
            figures.append(fig)

        pdf = FPDF()
        pdf.add_page()
        for fig_index , fig in enumerate(figures,1):
            img_path = f"plot_{fig_index}.png"
            fig.write_image(img_path)

            if pdf.get_y() + 100 > pdf.h :
                pdf.add_page()
            
            pdf.image(img_path , x=10 , y=pdf.get_y() + 30 , w=170)
            pdf.set_y(pdf.get_y() + 100)

            os.remove(img_path)

        pdf.output(filename,'F')

    def predict_formulas(self,selected_formulas_ids, nb_batches_to_remove , min_nb_batches, max_nb_adjustments, limit_initialization_date , weighted = False , initial_weight= 1 , increase_rate = 0.1):
        """
        Generate new formula reference recommendations based on specified grid search parameters.
        :param self: An instance of a Factory object.
        :type self: Factory
        :param selected_formulas_ids: A list of IDs representing the selected formulas.
        :type selected_formulas_ids: list
        :param nb_batches_to_remove: The number of batches to remove, typically those with the highest coefficient of variation.
        :type nb_batches_to_remove: int
        :param min_nb_batches: The minimum number of batches required for a formula to be included in the grid search process.
        :type min_nb_batches: int
        :param max_nb_adjustments: The maximum number of batch adjustments allowed for a formula to be included in the grid search process.
        :type max_nb_adjustments: int
        :param limit_initialization_date: The start date of the historical data to be considered in the grid search process.
        :type limit_initialization_date: pd.Period
        :param weighted: A boolean indicating whether to use a weighted average (giving more recent batches higher weight) for the recommendations.
        :type weighted: bool
        :param initial_weight: The initial weight assigned to the first batch in the selected history if 'weighted' is True.
        :type initial_weight: int
        :param increase_rate: The rate of increase in weight between consecutive batches if 'weighted' is True.
        :type increase_rate: int
        """
        filtered_formulas = {}
        filtered_formulas_removed = {}
        formulas_predictions = {}
        test_formulas = {}
        selected_batches_dict = {}
        formulas_ids_not_to_consider = []
        for selected_formula_id in selected_formulas_ids:
            selected_formula = self.formulas_dict[selected_formula_id]
            filtered_batches = []
            
            for batch in selected_formula.batches_arr[:-1]:
                if batch.initialization_date >= limit_initialization_date and batch.number_of_adjustement_operations <= max_nb_adjustments:
                    filtered_batches.append(batch)
            filtered_formulas[selected_formula_id] = filtered_batches
            test_formulas[selected_formula_id] = selected_formula.batches_arr[-1]
            

        for selected_formula_id in selected_formulas_ids:
            selected_batches = filtered_formulas[selected_formula_id]
            selected_formula = self.formulas_dict[selected_formula_id]
            if len(selected_batches) - nb_batches_to_remove < min_nb_batches :
                formulas_ids_not_to_consider.append(selected_formula_id) 
            else:
                test_formulas[selected_formula_id] = selected_formula.batches_arr[-1]
                indexes = selected_formula.calculate_sorted_indexes_by_squared_deviations(selected_batches)
                if nb_batches_to_remove > 0 :
                    indexes = indexes[:-nb_batches_to_remove]
                selected_batches_removed = [selected_batches[i] for i in indexes]
                filtered_formulas_removed[selected_formula_id] = selected_batches_removed

        selected_formulas_ids = [x for x in selected_formulas_ids if x not in formulas_ids_not_to_consider]
        
        for selected_formula_id in selected_formulas_ids:
            selected_batches = filtered_formulas_removed[selected_formula_id]
            selected_batches_dict[selected_formula_id] = selected_batches
            key_values = {}  # Dictionary to store values for each key across dictionaries

            # Populate key_values dictionary
            for batch in selected_batches:
                for key, value in batch.MP_total_relative_percentages.items():
                    if key not in key_values:
                        key_values[key] = []
                    key_values[key].append(value)


            # Calculate mean and variance for each key
            if weighted:
                weights = []
                for i in range(len(selected_batches)):
                    weight = initial_weight + i * increase_rate
                    weights.append(weight)
                sum_weights = sum(weights)
                #print(sum_weights)
                weighted_avg = {}
                for key, values in key_values.items():
                    sum_products = sum(value * weight for value,weight in zip(values,weights))
                    weighted_avg[key] = round(sum_products / sum_weights,4)

                formulas_predictions[selected_formula_id]=weighted_avg

            else:
                means = {}
                for key, values in key_values.items():
                    means[key] = np.mean(values)
                formulas_predictions[selected_formula_id]=means

        return formulas_predictions , test_formulas , selected_formulas_ids , formulas_ids_not_to_consider , selected_batches_dict

    def determine_metrics(self,selected_formulas_ids , formulas_predictions , test_formulas ):
        """
        Calculate metrics for the grid search process.
        :param self: An instance of a Factory object.
        :type self: Factory
        :param selected_formulas_ids: A list containing the IDs of the selected formulas.
        :type selected_formulas_ids: list
        :param formulas_predictions: A dictionary containing the recommended new formula references.
        :type formulas_predictions: dict
        :param test_formulas: A dictionary representing the formula references used for testing in the grid search process.
        :type test_formulas: dict
        """
        metrics_MP = {}
        metrics_formulas = {}
        for selected_formula_id in selected_formulas_ids:
            formula_prediction = formulas_predictions[selected_formula_id]
            test_formula = test_formulas[selected_formula_id]
            metric = {key: ((formula_prediction[key] - test_formula.MP_total_relative_percentages[key]) / test_formula.MP_total_relative_percentages[key]) * 100 for key in test_formula.MP_total_relative_percentages.keys()}
            metrics_MP[selected_formula_id] = metric
            absolute_values = [abs(value) for value in metric.values()]
            mean_absolute_values = sum(absolute_values)/ len(absolute_values)
            metrics_formulas[selected_formula_id] = mean_absolute_values
        
        return metrics_MP , metrics_formulas
    
    def function_metric(self, selected_formulas_ids , nb_batches_to_remove , min_nb_batches, max_nb_adjustments, limit_initialization_date , weighted = False , initial_weight= 1 , increase_rate = 0.1):
        """
        Generate new formula reference recommendations and detemine the metrics based on specified grid search parameters.
        :param self: An instance of a Factory object.
        :type self: Factory
        :param selected_formulas_ids: A list of IDs representing the selected formulas.
        :type selected_formulas_ids: list
        :param nb_batches_to_remove: The number of batches to remove, typically those with the highest coefficient of variation.
        :type nb_batches_to_remove: int
        :param min_nb_batches: The minimum number of batches required for a formula to be included in the grid search process.
        :type min_nb_batches: int
        :param max_nb_adjustments: The maximum number of batch adjustments allowed for a formula to be included in the grid search process.
        :type max_nb_adjustments: int
        :param limit_initialization_date: The start date of the historical data to be considered in the grid search process.
        :type limit_initialization_date: pd.Period
        :param weighted: A boolean indicating whether to use a weighted average (giving more recent batches higher weight) for the recommendations.
        :type weighted: bool
        :param initial_weight: The initial weight assigned to the first batch in the selected history if 'weighted' is True.
        :type initial_weight: int
        :param increase_rate: The rate of increase in weight between consecutive batches if 'weighted' is True.
        :type increase_rate: int
        """
        formulas_predictions , test_formulas , filtered_formulas_ids , formulas_ids_filtered_out , selected_batches_dict= self.predict_formulas(selected_formulas_ids, nb_batches_to_remove , min_nb_batches, max_nb_adjustments, limit_initialization_date , weighted , initial_weight , increase_rate)
        metrics_MP , metrics_formulas = self.determine_metrics(filtered_formulas_ids , formulas_predictions , test_formulas)
        if len(metrics_formulas) == 0:
            average_metric = float('inf')
        else:
            average_metric = round(sum(metrics_formulas.values())/len(metrics_formulas),5)

        percentage_of_formulas_filtered_out = round(len(formulas_ids_filtered_out)/len(selected_formulas_ids) , 3)*100
        return average_metric , percentage_of_formulas_filtered_out , formulas_predictions , metrics_MP , formulas_ids_filtered_out , selected_batches_dict

    def grid_search(self ,selected_formulas_ids ,percentage_filtered_out_formulas_threshold, func, param_ranges ):
        """
        Execute the grid search process.
        :param self: An instance of a Factory object.
        :type self: Factory
        :param selected_formulas_ids: A list containing the IDs of the selected formulas.
        :type selected_formulas_ids: list
        :param percentage_filtered_out_formulas_threshold: The maximum allowed percentage of formulas that can be filtered out during the grid search process after the filtering stage.
        :type percentage_filtered_out_formulas_threshold: int
        :param func: The function to be invoked during the grid search process.
        :type func: function
        :param param_ranges: A dictionary containing the parameter ranges to be passed to the selected function.
        :type param_ranges: dict
        """
        # Get all combinations of parameter values
        param_combinations = product(*param_ranges.values())

        # Initialize variables to store best parameters and scores
        best_params = None
        best_metrics = float('inf'),float('inf'),float('inf'),float('inf'),float('inf') # Initialize with very large values

        entered = False
        # Perform grid search
        i=0
        for params in tqdm(param_combinations, desc="Processing search"):
            i = i+1
            if  params[2]  < 2 :
                break
            metrics = func(selected_formulas_ids, *params)
            # Check if the current metrics are better than the best metrics found so far
            if metrics[0] < best_metrics[0] and metrics[1] < percentage_filtered_out_formulas_threshold:
                entered = True
                best_metrics = metrics
                best_params = params
                selected_batches_dict = metrics[-1]
        if entered :
            return True , best_params, best_metrics[:2] , best_metrics[2] , best_metrics[3] , best_metrics[4] , selected_batches_dict
        else:
            return False , [] , [] , [] , [] , [] ,  []

    def search_of_stable_formulas(self ,selected_formulas_ids , threshold_adjustements_accepted ,nb_of_recent_batches_considered , threshold_cv , min_accepted_batches):
        """
        Identify a list of stable formulas based on the average raw material composition.
        :param self: An instance of a Factory object.
        :type self: Factory
        :param selected_formulas_ids: A list containing the IDs of the selected formulas.
        :type selected_formulas_ids: list
        :param threshold_adjustements_accepted: The maximum number of batch adjustments allowed for a formula to be considered stable.
        :type threshold_adjustements_accepted: int
        :param nb_of_recent_batches_considered: The number of recent batches to consider when assessing the stability of a formula.
        :type nb_of_recent_batches_considered: int
        :param threshold_cv: The maximum acceptable coefficient of variation for a formula to be considered stable.
        :type threshold_cv: float
        :param min_accepted_batches: The minimum number of batches required for a formula to be considered stable.
        :type min_accepted_batches: int
        """
        stable_formulas = []
        for formula_id in selected_formulas_ids:
            is_stable = True
            selected_formula = self.formulas_dict[formula_id]
            #print(selected_formula)
            selected_stable_batches = []
            for selected_batch in selected_formula.batches_arr[-nb_of_recent_batches_considered:]:
                if selected_batch.number_of_adjustement_operations <= threshold_adjustements_accepted :
                    selected_stable_batches.append(selected_batch)

            key_values = {}  # Dictionary to store values for each key across dictionaries

            if len(selected_stable_batches) >= min_accepted_batches :
                # Populate key_values dictionary
                for batch in selected_stable_batches:
                    for key, value in batch.MP_total_relative_percentages.items():
                        if key not in key_values:
                            key_values[key] = []
                        key_values[key].append(value)

                # Calculate mean and variance for each key
                means = {}
                variances = {}
                for key, values in key_values.items():
                    means[key] = np.mean(values)
                    variances[key] = np.var(values)

                # Calculate coefficient of variation (CV)
                cvs = {}
                for key, mean in means.items():
                    variance = variances[key]
                    cv = np.sqrt(variance) / mean
                    cvs[key] = cv

                for cv in cvs.values():
                    if cv > threshold_cv :
                        is_stable = False
            else:
                is_stable = False
        
            if is_stable : 
                stable_formulas.append({formula_id : means})
        return stable_formulas
          

    def export_predicted_and_ref_formulas_to_excel(self,filename ,threshold_percentage , formulas_predictions , formulas_ref_df , four_ten_days_requested_formulas_df_grouped , four_days_one_month_requested_formulas_df_grouped ,earliest_date_after_four_days,  formulas_ids_filtered_out, stable_formulas_ids_list , stable_last_batch_formulas_ids_list, parameters_metrics_variables):  
        """
        Export the results to an Excel file.
        :param self: An instance of a Factory object.
        :type self: Factory
        :param filename: The name of the Excel file to save the results.
        :type filename: str
        :param threshold_percentage: The threshold percentage for variation between the predicted relative percentage and the reference relative percentage of each raw material. Cells with a variation higher than this threshold will be highlighted in red.
        :type threshold_percentage: int
        :param formulas_predictions: A DataFrame containing the new predicted references for the formulas.
        :type formulas_predictions: pd.DataFrame
        :param formulas_ref_df: A DataFrame containing the reference data for the formulas.
        :type formulas_ref_df: pd.DataFrame
        :param four_ten_days_requested_formulas_df_grouped: A DataFrame containing the number of requested batches for each formula within a four to ten days interval after the specified date.
        :type four_ten_days_requested_formulas_df_grouped: pd.DataFrame
        :param four_days_one_month_requested_formulas_df_grouped: A DataFrame containing the number of requested batches for each formula within a four days to one month interval after the specified date.
        :type four_days_one_month_requested_formulas_df_grouped: pd.DataFrame
        :param earliest_date_after_four_days: The earliest date for a requested batch, starting from four days after the specified date.
        :type earliest_date_after_four_days: str
        :param formulas_ids_filtered_out: A list of formula IDs that have been filtered out from the grid search process.
        :type formulas_ids_filtered_out: list
        :param stable_formulas_ids_list: A list of formula IDs considered to be stable.
        :type stable_formulas_ids_list: list
        :param stable_last_batch_formulas_ids_list: A list of formula IDs that have stable last batches with no adjustments.
        :type stable_last_batch_formulas_ids_list: list
        :param parameters_metrics_variables: A dictionary containing the parameters and main metrics for the stable formulas filtering and grid search processes.
        :type parameters_metrics_variables: dict
        """

        with pd.ExcelWriter(filename) as writer :
            # Create a workbook object
            workbook = writer.book

            # Add a worksheet to the workbook
            formulas_predictions_worksheet = workbook.add_worksheet('formulas predictions')
            formulas_indicators_worksheet = workbook.add_worksheet('formulas indicators')
            parameters_metrics_worksheet = workbook.add_worksheet('parameters and metrics')

            formulas_predictions_title_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter'})
            formulas_indicators_titles_list = ['N° formule','Description','Type','Stable','Last_batch_stable','Filtered_out','Besoin 4/10 jours (nb batchs)','Besoin 4 jours à 1 mois (nb batchs)','Date prochain besoin','Recommendations summary','Nb ajustements moyen (average of last 3 batches)','Nb adjustments for last batch','Ampleur modification formule (RM > 2%)' , 'Dates of latest modification of formula']
            for i,title in enumerate(formulas_indicators_titles_list):
                formulas_indicators_worksheet.write(0,i,title,formulas_predictions_title_format)
            # Define formats for title and dataframe
            formulas_predictions_title_format = workbook.add_format({'bold': True, 'border': 2, 'align': 'center', 'valign': 'vcenter'})
            dataframe_format = workbook.add_format({'border': 1})
            # Write each DataFrame to the Excel sheet with specific formatting
            formulas_indicators_start_row = 1  # Starting row for titles
            formulas_predictions_start_row = 1  # Starting row for titles
            parameters_metrics_worksheet_start_row = 1

            formula_ids_list = formulas_predictions["productno"].unique().tolist()
            for formula_id in formula_ids_list :
                    # Convert dictionaries to pandas dataframes
                    formulas_predictions_df = formulas_predictions[formulas_predictions['productno'] == formula_id]
                    formulas_ref_df_aux = formulas_ref_df[formulas_ref_df['productno'] == formula_id]
                    # Merge dataframes on 'Keys' column
                    merged_df = pd.merge(formulas_predictions_df[['MP','formulas_predictions']], formulas_ref_df_aux[['Produit','formulas_ref']] ,left_on="MP",right_on='Produit')
                    merged_df = merged_df.drop(columns=['Produit'])
                    merged_df['formulas_predictions'] = merged_df['formulas_predictions'].round(4)
                    merged_df['difference_percentage_to_ref'] = round(((merged_df['formulas_predictions'] - merged_df['formulas_ref'])/merged_df['formulas_ref'])*100,1)
                    
                    # Apply the style
                    styled_df = merged_df.style.apply(highlight_above_threshold, subset=['difference_percentage_to_ref'], threshold=threshold_percentage)
                    is_above = np.abs(merged_df['difference_percentage_to_ref']) > threshold_percentage
                    nb_of_above_threshold = is_above.sum()

                    # Write title above the DataFrame
                    formulas_predictions_title = f'N° formule : {formula_id} , Description : {self.formulas_dict[formula_id].description}'

                    formulas_predictions_worksheet.merge_range(formulas_predictions_start_row, 1, formulas_predictions_start_row, merged_df.shape[1], formulas_predictions_title, formulas_predictions_title_format)


                    formulas_indicators_worksheet.write(formulas_indicators_start_row,0,formula_id)
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,1,self.formulas_dict[formula_id].description)
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,2,self.formulas_dict[formula_id].type)
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,3,False)
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,4,False)
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,5,False)

                    row_index = four_ten_days_requested_formulas_df_grouped[four_ten_days_requested_formulas_df_grouped['OP_ARTICLE']==formula_id]
                    if not row_index.empty :
                        formulas_indicators_worksheet.write(formulas_indicators_start_row,6,four_ten_days_requested_formulas_df_grouped.at[row_index.index[0],'OP_ORDRE'])
                    else:
                        formulas_indicators_worksheet.write(formulas_indicators_start_row,6,0)

                    row_index = four_days_one_month_requested_formulas_df_grouped[four_days_one_month_requested_formulas_df_grouped['OP_ARTICLE']==formula_id]

                    if  not row_index.empty :
                        formulas_indicators_worksheet.write(formulas_indicators_start_row,7,four_days_one_month_requested_formulas_df_grouped.at[row_index.index[0],'OP_ORDRE'])
                    else:
                        formulas_indicators_worksheet.write(formulas_indicators_start_row,7,0)

                    row_index = earliest_date_after_four_days[earliest_date_after_four_days['OP_ARTICLE']==formula_id]

                    if  not row_index.empty :
                        formulas_indicators_worksheet.write(formulas_indicators_start_row,8,earliest_date_after_four_days.at[row_index.index[0],'OP_DATE_DEBUT'])
                    else:
                        formulas_indicators_worksheet.write(formulas_indicators_start_row,8,"-")
                    
                    predictions_references_df = merged_df
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,9,generate_recommendation_summary(predictions_references_df , threshold_percentage))
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,10,self.formulas_dict[formula_id].average_number_of_adjustements)
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,11,self.formulas_dict[formula_id].batches_arr[-1].number_of_adjustement_operations)
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,12,nb_of_above_threshold)
                    if not self.formulas_dict[formula_id].list_of_dates_of_modification_of_formulas:
                        formulas_indicators_worksheet.write(formulas_indicators_start_row,13,"-")
                    else:
                        formulas_indicators_worksheet.write(formulas_indicators_start_row,13,str(self.formulas_dict[formula_id].list_of_dates_of_modification_of_formulas[-1]))
                    


                   # Write the DataFrame to the Excel sheet
                    styled_df.to_excel(writer, sheet_name='formulas predictions', startrow=formulas_predictions_start_row + 1, startcol=1, index=False)

                    # Update start_row for the next title
                    formulas_indicators_start_row += 1 # Add 2 for the space and 1 for the title
                    formulas_predictions_start_row += merged_df.shape[0] + 4


            for formula_id in formulas_ids_filtered_out :

                selected_formula = self.formulas_dict[formula_id]

                formulas_indicators_worksheet.write(formulas_indicators_start_row,0,formula_id)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,1,selected_formula.description)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,2,self.formulas_dict[formula_id].type)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,3,False)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,4,False)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,5,True)


                row_index = four_ten_days_requested_formulas_df_grouped[four_ten_days_requested_formulas_df_grouped['OP_ARTICLE']==formula_id]
                if not row_index.empty :
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,6,four_ten_days_requested_formulas_df_grouped.at[row_index.index[0],'OP_ORDRE'])
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,6,0)

                row_index = four_days_one_month_requested_formulas_df_grouped[four_days_one_month_requested_formulas_df_grouped['OP_ARTICLE']==formula_id]

                if  not row_index.empty :
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,7,four_days_one_month_requested_formulas_df_grouped.at[row_index.index[0],'OP_ORDRE'])
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,7,0)

                row_index = earliest_date_after_four_days[earliest_date_after_four_days['OP_ARTICLE']==formula_id]

                if  not row_index.empty :
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,8,earliest_date_after_four_days.at[row_index.index[0],'OP_DATE_DEBUT'])
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,8,"-")
                    
                formulas_indicators_worksheet.write(formulas_indicators_start_row,9,"-")
                formulas_indicators_worksheet.write(formulas_indicators_start_row,10,selected_formula.average_number_of_adjustements)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,11,selected_formula.batches_arr[-1].number_of_adjustement_operations)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,12,"-")
                if not self.formulas_dict[formula_id].list_of_dates_of_modification_of_formulas:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,13,"-")
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,13,str(self.formulas_dict[formula_id].list_of_dates_of_modification_of_formulas[-1]))
                formulas_indicators_start_row += 1  # Add 2 for the space and 1 for the title

            for formula_id in stable_formulas_ids_list :

                selected_formula = self.formulas_dict[formula_id]

                formulas_indicators_worksheet.write(formulas_indicators_start_row,0,formula_id)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,1,selected_formula.description)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,2,self.formulas_dict[formula_id].type)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,3,True)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,4,True)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,5,"-")
                row_index = four_ten_days_requested_formulas_df_grouped[four_ten_days_requested_formulas_df_grouped['OP_ARTICLE']==formula_id]
                if not row_index.empty :
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,6,four_ten_days_requested_formulas_df_grouped.at[row_index.index[0],'OP_ORDRE'])
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,6,0)

                row_index = four_days_one_month_requested_formulas_df_grouped[four_days_one_month_requested_formulas_df_grouped['OP_ARTICLE']==formula_id]

                if  not row_index.empty :
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,7,four_days_one_month_requested_formulas_df_grouped.at[row_index.index[0],'OP_ORDRE'])
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,7,0)

                row_index = earliest_date_after_four_days[earliest_date_after_four_days['OP_ARTICLE']==formula_id]

                if  not row_index.empty :
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,8,earliest_date_after_four_days.at[row_index.index[0],'OP_DATE_DEBUT'])
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,8,"-")
                
                formulas_indicators_worksheet.write(formulas_indicators_start_row,9,"-")
                formulas_indicators_worksheet.write(formulas_indicators_start_row,10,selected_formula.average_number_of_adjustements)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,11,selected_formula.batches_arr[-1].number_of_adjustement_operations)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,12,"-")
                if not self.formulas_dict[formula_id].list_of_dates_of_modification_of_formulas:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,13,"-")
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,13,str(self.formulas_dict[formula_id].list_of_dates_of_modification_of_formulas[-1]))
                formulas_indicators_start_row += 1  # Add 2 for the space and 1 for the title

            for formula_id in stable_last_batch_formulas_ids_list :

                selected_formula = self.formulas_dict[formula_id]

                formulas_indicators_worksheet.write(formulas_indicators_start_row,0,formula_id)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,1,selected_formula.description)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,2,self.formulas_dict[formula_id].type)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,3,False)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,4,True)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,5,"-")
                row_index = four_ten_days_requested_formulas_df_grouped[four_ten_days_requested_formulas_df_grouped['OP_ARTICLE']==formula_id]
                if not row_index.empty :
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,6,four_ten_days_requested_formulas_df_grouped.at[row_index.index[0],'OP_ORDRE'])
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,6,0)

                row_index = four_days_one_month_requested_formulas_df_grouped[four_days_one_month_requested_formulas_df_grouped['OP_ARTICLE']==formula_id]

                if  not row_index.empty :
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,7,four_days_one_month_requested_formulas_df_grouped.at[row_index.index[0],'OP_ORDRE'])
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,7,0)

                row_index = earliest_date_after_four_days[earliest_date_after_four_days['OP_ARTICLE']==formula_id]

                if  not row_index.empty :
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,8,earliest_date_after_four_days.at[row_index.index[0],'OP_DATE_DEBUT'])
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,8,"-")
                
                formulas_indicators_worksheet.write(formulas_indicators_start_row,9,"-")
                formulas_indicators_worksheet.write(formulas_indicators_start_row,10,selected_formula.average_number_of_adjustements)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,11,selected_formula.batches_arr[-1].number_of_adjustement_operations)
                formulas_indicators_worksheet.write(formulas_indicators_start_row,12,"-")
                if not self.formulas_dict[formula_id].list_of_dates_of_modification_of_formulas:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,13,"-")
                else:
                    formulas_indicators_worksheet.write(formulas_indicators_start_row,13,str(self.formulas_dict[formula_id].list_of_dates_of_modification_of_formulas[-1]))
                formulas_indicators_start_row += 1  # Add 2 for the space and 1 for the title

        path = project_directory + "/excel templates/parameters_and_metrics template.xlsx"

        src_wb = openpyxl.load_workbook(path)
        src_ws = src_wb['parameters and metrics']
        dest_wb = openpyxl.load_workbook(filename)
        dest_ws = dest_wb['parameters and metrics']
        max_row = src_ws.max_row
        max_col = src_ws.max_column
        
        for row in src_ws.iter_rows(min_row=1 , max_row=max_row , min_col=1 , max_col=max_col):
            for cell in row:
                new_cell = dest_ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value  # Set new value
                copy_formatting(cell, new_cell)  # Copy formatting

        # Update cells with variables
        if parameters_metrics_variables:
            for key, value in parameters_metrics_variables.items():
                dest_ws[key].value = value

        # Save the destination workbook
        dest_wb.save(filename)
        print("Results excel file is created .")


class Formula:
    def __init__(self,formula_df):
        """
        Intiate a formula object instance from the factory dataframe.
        :param self: a formula object instance
        :type self: Formula object
        :param formula_df: a dataframe of the formula 
        :type formula_df: pd.Dataframe
        """
        formula_df_sorted = formula_df.sort_values(by='creation_date', ascending=True)
        self.formula_id = (formula_df['N° formule'].tolist())[0]
        self.description = (formula_df['Description'].tolist())[0]
        self.type = (formula_df['Type'].tolist())[0]
        self.Besoins_7_jours = (formula_df['Besoins 7 jours'].tolist())[0]
        self.Besoins_1_mois = (formula_df['Besoins 1 mois'].tolist())[0]
        self.MP_set = determine_MP_set(formula_df_sorted)
        self.batches_arr = []
        self.ingest_formula_data(formula_df_sorted)
        self.formula_reference = self.batches_arr[0]
        self.rest_batches_arr = self.batches_arr[1:]
        self.number_of_adjusted_batches = self.determine_number_of_adjusted_batches()
        self.average_number_of_adjustements = self.determine_number_of_adjustement_operations_average()
        self.percentage_of_adjusted_batches = round((self.number_of_adjusted_batches / len(self.batches_arr))*100 ,2)
        self.coefficient_of_variation_of_MP = self.determine_coefficient_of_variation_for_MP()
        self.list_of_dates_of_modification_of_formulas , self.no_unique_raw_material_composition = self.determine_list_of_dates_of_modification_of_formulas(threshold_accepted_modification = 0.01)

    def ingest_formula_data(self, formula_df_sorted):
        """
        Ingest a formula data from the factory dataframe.
        :param self: a formula object
        :type self: Formula object
        :param formula_df_sorted: a dataframe of the formula (with batches sorted on production time)
        :type formula_df_sorted: pd.Dataframe
        """
        batches_set = formula_df_sorted['WipOrderNo'].unique().tolist()
        for batch in batches_set:
            batch_df = formula_df_sorted[formula_df_sorted['WipOrderNo'] == batch]
            batch_created = Batch(batch_df = batch_df , MP_set=self.MP_set)
            self.batches_arr.append(batch_created)

    def determine_list_of_dates_of_modification_of_formulas(self , threshold_accepted_modification):
        """
        Determine a list of dates corresponding for the modifications of the formula.
        :param self: a formula object
        :type self: Formula object
        :param threshold_accepted_modification: a threshold for considering a modification
        :type threshold_accepted_modification: float
        """
        no_unique_raw_material_composition = False
        list_of_dates_of_modification_of_formulas = []  # Dictionary to store values for each key across dictionaries
        for i in range(len(self.batches_arr)-1) :
            old_batch , new_batch = self.batches_arr[i] , self.batches_arr[i+1]
            relative_errors = {}
            for key in old_batch.MP_initial_relative_percentages.keys():
                if old_batch.MP_initial_relative_percentages[key] == 0:
                    no_unique_raw_material_composition= True
                    relative_errors[key] = threshold_accepted_modification+1
                else:
                    error = abs(old_batch.MP_initial_relative_percentages[key] - new_batch.MP_initial_relative_percentages[key])/ abs(old_batch.MP_initial_relative_percentages[key])
                    relative_errors[key] = error
            #average_relative_error = sum(relative_errors.values()) / len(relative_errors.values())
            for relative_error in relative_errors.values():
                if (relative_error >= threshold_accepted_modification) :
                    list_of_dates_of_modification_of_formulas.append(new_batch.initialization_date)
                    break

        
        return list(set(list_of_dates_of_modification_of_formulas)) , no_unique_raw_material_composition

    def determine_number_of_adjusted_batches(self):
        """
        Determine number of adjusted batches of a formula.
        :param self: a formula object
        :type self: Formula object
        """
        number_of_adjusted_batches = 0
        for i,batch in enumerate(self.batches_arr):
            if batch.adjusted :
                number_of_adjusted_batches = number_of_adjusted_batches + 1
        return number_of_adjusted_batches
    
    def determine_number_of_adjustement_operations_average(self):
        """
        Determine average number of adjustements of a formula.
        :param self: a formula object
        :type self: Formula object
        """
        number_of_adjustement_operations_average = 0
        for i,batch in enumerate(self.batches_arr):
            if len(self.batches_arr)<3:
                number_of_adjustement_operations_average += batch.number_of_adjustement_operations
            if len(self.batches_arr)>=3 and i>= len(self.batches_arr)-3:
                number_of_adjustement_operations_average += batch.number_of_adjustement_operations
        if len(self.batches_arr)<3:
            number_of_adjustement_operations_average = int(number_of_adjustement_operations_average/ (len(self.batches_arr)))
        else:
            number_of_adjustement_operations_average = int(number_of_adjustement_operations_average/ 3)
        return number_of_adjustement_operations_average
    
    def determine_coefficient_of_variation_for_MP(self):
        """
        Determine coefficient of variation (CV) for each raw material (MP) of the formula.
        :param self: a formula object
        :type self: Formula object
        """
        key_values = {}  # Dictionary to store values for each key across dictionaries
        cvs = {}

        for batch in self.batches_arr:
            for key, value in batch.MP_total_relative_percentages.items():
                if key not in key_values:
                    key_values[key] = []
                key_values[key].append(value)

        # Calculate coefficient of variation (CV) for each key
        for key, values in key_values.items():
            mean = np.mean(values)
            variance = np.var(values)
            cv = np.sqrt(variance) / mean
            cvs[key] = cv

        return cvs
    
    def calculate_sorted_indexes_by_squared_deviations(self , batches_arr):
        """
        Return indexes of batches of the formula sorted (by ascending order) by summed squared deviations of raw materials of each batch .
        :param self: a formula object
        :type self: Formula object
        :param batches_arr: a list of batches for selected formula
        :type batches_arr: list of Batch object
        """
        key_values = {}  # Dictionary to store values for each key across dictionaries

        # Populate key_values dictionary
        for batch in batches_arr:
            for key, value in batch.MP_total_relative_percentages.items():
                if key not in key_values:
                    key_values[key] = []
                key_values[key].append(value)

        # Calculate mean and variance for each key
        squared_deviations = {}  # Store squared deviations
        for key, values in key_values.items():
            mean = np.mean(values)
            # Calculate squared deviations
            squared_deviations[key] = [np.abs((value - mean)/value) for value in values]

        # Sum squared deviations for each batch across all keys
        batch_sum_squared_deviations = [sum(batch_deviation) for batch_deviation in zip(*squared_deviations.values())]
        
        # Sort batches based on summed squared deviations
        sorted_indexes = sorted(range(len(batch_sum_squared_deviations)), key=lambda k: batch_sum_squared_deviations[k])

        return sorted_indexes

    def plot_bar_graph(self):
        """
        Plot a bar graph for selected formula .
        :param self: a formula object
        :type self: Formula object
        """
        number_of_adjustements_selected = []
        batches_ids = []
        for batch in self.batches_arr:
            number_of_adjustements_selected.append(batch.number_of_adjustement_operations)
            batches_ids.append(batch.batch_id)

        # Initialize empty lists for each key
        keys = self.batches_arr[0].MP_total_relative_percentages.keys()
        key_lists = {key: [] for key in keys}

        # Iterate through the list of dictionaries
        for batch in self.batches_arr:
            # Iterate through each key in the dictionary
            for key, value in batch.MP_total_relative_percentages.items():
            # Append the value to the corresponding key list
                key_lists[key].append(value)

        fig = make_subplots(rows=1, cols=2, shared_xaxes=True, subplot_titles=("Relative percentage of each MP", "Number of adjustements applied"))

        # Plotly bar plot for the primary y-axis
        for column in key_lists:  # Exclude the first column which is the x-axis values
            fig.add_trace(go.Bar(x=batches_ids, y=key_lists[column], name=column), row=1, col=1)

        # Plotly bar plot for the secondary y-axis
        fig.add_trace(go.Bar(x=batches_ids, y=number_of_adjustements_selected, name='Count of adjustements'), row=1, col=2)

        fig.update_xaxes(title_text='WipOrderNo', row=1, col=1, type='category')
        fig.update_xaxes(title_text='WipOrderNo', row=1, col=2, type='category',)

        fig.update_layout(
            title=f'Relative percentage of MP adjusted in orders for {self.formula_id} \n Percentage of adjusted orders : {self.percentage_of_adjusted_batches} %',
            xaxis=dict(
                title='WipOrderNo',
                type='category',
                categoryorder='array',
                categoryarray=batches_ids  # Ensure correct order of categories
            ),
            legend_title='ComponentProductNo',
            barmode='stack',
            width=1000,
            height=500
        )
        # Set the same category order for both subplots
        fig.update_xaxes(categoryorder='array', categoryarray=batches_ids, row=1, col=1)
        fig.update_xaxes(categoryorder='array', categoryarray=batches_ids, row=1, col=2)

        return fig
    
    """TODO correct the mistake of plotting base for 'plot_bar_graph_with_ref' function"""
    def plot_bar_graph_with_ref(self):
        """
        Plot a bar graph of ratio of each raw material (MP) with respect to reference batch for selected formula .
        :param self: a formula object
        :type self: Formula object
        """
        number_of_adjustements_selected = []
        batches_ids = []
        for batch in self.rest_batches_arr:
            number_of_adjustements_selected.append(batch.number_of_adjustement_operations)
            batches_ids.append(batch.batch_id)


        # Initialize empty lists for each key
        keys = self.rest_batches_arr[0].MP_total_relative_percentages.keys()
        key_lists = {key: [] for key in keys}

        # Iterate through the list of dictionaries
        for batch in self.rest_batches_arr:
            # Iterate through each key in the dictionary
            for key, value in batch.MP_total_relative_percentages.items():
            # Append the value to the corresponding key list
                val = (value / self.formula_reference.MP_total_relative_percentages[key])*100 - 100
                key_lists[key].append(val)

        fig = make_subplots(rows=1, cols=2, shared_xaxes=True, subplot_titles=("Ratio of each MP with respect to ref", "Number of adjustements applied"))

        # Plotly bar plot for the primary y-axis
        for column in key_lists:  # Exclude the first column which is the x-axis values
            fig.add_trace(go.Bar(x=batches_ids, y=key_lists[column], name=column), row=1, col=1)

        # Plotly bar plot for the secondary y-axis
        fig.add_trace(go.Bar(x=batches_ids, y=number_of_adjustements_selected, name='Count of adjustements'), row=1, col=2)

        fig.update_xaxes(title_text='WipOrderNo', row=1, col=1, type='category')
        fig.update_xaxes(title_text='WipOrderNo', row=1, col=2, type='category',)

        fig.update_layout(
            title=f'Relative percentage of MP adjusted in orders for {self.formula_id} \n Percentage of adjusted orders : {self.percentage_of_adjusted_batches} %',
            xaxis=dict(
                title='WipOrderNo',
                type='category',
                categoryorder='array',
                categoryarray=batches_ids  
            ),
            legend_title = 'ComponentProductNo',
            barmode='stack',
            width=1000,
            height=500
        )
        # Set the same category order for both subplots
        fig.update_xaxes(categoryorder='array', categoryarray=batches_ids, row=1, col=1)
        fig.update_xaxes(categoryorder='array', categoryarray=batches_ids, row=1, col=2)

        return fig


class Batch:
    def __init__(self , batch_df , MP_set):
        """
        Intiate a batch object instance from the factory dataframe.
        :param self: a batch object instance
        :type self: Batch object
        :param batch_df: a dataframe of the batch 
        :type batch_df: pd.Dataframe
        :param MP_set: a set of raw material of 
        :type MP_set: set of raw materials of formula to which selected batch belong
        """
        self.batch_id = (batch_df['WipOrderNo'].tolist())[0]
        self.adjusted = True if (batch_df[batch_df['Rectif ou Ajust Pesée en Gr'] > 0.0]).shape[0] > 0 else False
        self.initialization_date = (batch_df['creation_date'].tolist())[0]

        self.MP_initial_amounts = {key: 0 for key in MP_set}
        self.MP_adjusted_amounts = {key: 0 for key in MP_set}
        self.MP_total_amounts = {key: 0 for key in MP_set}
        self.MP_initial_relative_percentages = {key: 0 for key in MP_set}
        self.MP_adjusted_relative_percentages = {key: 0 for key in MP_set}
        self.MP_total_relative_percentages = {key: 0 for key in MP_set}

        self.ingest_batch_data(batch_df ,MP_set)
        self.number_of_adjustement_operations = self.count_of_adjustement_operations(batch_df)
   
    def ingest_batch_data(self, batch_df , MP_set):
        """
        Ingest a batch data from the factory dataframe.
        :param self: a batch object instance
        :type self: Batch object
        :param batch_df: a dataframe of the batch 
        :type batch_df: pd.Dataframe
        :param MP_set: a set of raw material of 
        :type MP_set: set of raw materials of formula to which selected batch belong
        """
        for MP_selected in MP_set : 
            batch_df_grouped =batch_df.groupby(['ComponentProductNo']).agg({'Intro Pesée en Gr': 'sum', 'Rectif ou Ajust Pesée en Gr': 'sum' , 'IntroKGM_total': 'sum'}).reset_index()
            if len(batch_df_grouped[batch_df_grouped['ComponentProductNo']==MP_selected]['Intro Pesée en Gr']) != 0 : 
                self.MP_initial_amounts[MP_selected] = batch_df_grouped[batch_df_grouped['ComponentProductNo']==MP_selected]['Intro Pesée en Gr'].iloc[0]
            if len(batch_df_grouped[batch_df_grouped['ComponentProductNo']==MP_selected]['Rectif ou Ajust Pesée en Gr']) != 0 :
                self.MP_adjusted_amounts[MP_selected] = batch_df_grouped[batch_df_grouped['ComponentProductNo']==MP_selected]['Rectif ou Ajust Pesée en Gr'].iloc[0]
            if len(batch_df_grouped[batch_df_grouped['ComponentProductNo']==MP_selected]['IntroKGM_total']) != 0 :
                self.MP_total_amounts[MP_selected] = batch_df_grouped[batch_df_grouped['ComponentProductNo']==MP_selected]['IntroKGM_total'].iloc[0]

        for MP_selected in MP_set :
            self.MP_initial_relative_percentages[MP_selected] = (self.MP_initial_amounts[MP_selected] / sum(self.MP_initial_amounts.values()))*100
            # Calculate the sum of adjusted amounts excluding NaN and infinite values
            adjusted_sum = sum(self.MP_adjusted_amounts.values())

            # Check if the sum is not zero
            if adjusted_sum != 0:
            # Calculate adjusted relative percentages
                self.MP_adjusted_relative_percentages[MP_selected] = (self.MP_adjusted_amounts[MP_selected] / adjusted_sum) * 100
            else:
            # Handle the case where the sum is zero
                self.MP_adjusted_relative_percentages[MP_selected] = 0
            self.MP_total_relative_percentages[MP_selected] = (self.MP_total_amounts[MP_selected] / sum(self.MP_total_amounts.values()))*100

    def count_of_adjustement_operations(self, batch_df):
        """
        Return number of adjustements done for a batch .
        :param self: a batch object instance
        :type self: Batch object
        :param batch_df: a dataframe of the batch 
        :type batch_df: pd.Dataframe
        """
        if self.adjusted:
            #Convert 'datetime_column' to datetime if it's not already in datetime format
            batch_df_sorted = batch_df.sort_values(by='Date Intro', ascending=True)

            # Calculate the slot for each datetime value
            batch_df_sorted['slot'] = (batch_df_sorted['Date Intro'].dt.hour * 60 + batch_df_sorted['Date Intro'].dt.minute) // 30

             # Convert back to datetime format for the slot
            batch_df_sorted['slot_datetime'] = pd.to_datetime(batch_df_sorted['Date Intro'].dt.date) + pd.to_timedelta(batch_df_sorted['slot'] * 30, unit='minutes')

            batch_df_grouped = batch_df_sorted.groupby(['slot_datetime']).agg({'Date Intro': 'count'}).reset_index()
            number_of_adjustements = batch_df_grouped.shape[0]
            return number_of_adjustements
        return 0